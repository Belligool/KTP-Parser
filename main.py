import os
import time
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from core.ocr_engine import OCREngine
from core.extractor import KTPExtractor

def main():
    start_time = time.perf_counter()
    target_dir = 'tests/sample_pdfs/'
    
    if not os.path.exists(target_dir):
        print(f"Directory '{target_dir}' not found. Please create it and add KTP PDFs.")
        return
    
    ocr = OCREngine()
    extractor = KTPExtractor()
    
    all_ktp_records = []
    print(f"Starting batch KTP extraction from {target_dir}...\n")
    
    for filename in os.listdir(target_dir):
        if filename.lower().endswith('.pdf'):
            pdf_path = os.path.join(target_dir, filename)
            print(f"Scanning: {filename}...")
            raw_text, low_confidence_words = ocr.extract(pdf_path)
            
            if raw_text:
                parsed_data = extractor.extract_data(raw_text, low_confidence_words)
                parsed_data['Source File'] = filename
                all_ktp_records.append(parsed_data)
            else:
                print(f"  -> WARNING: No text could be extracted from {filename}")
                
    if all_ktp_records:
        output_dir = 'tests/results/'
        os.makedirs(output_dir, exist_ok=True)
        
        excel_filename = os.path.join(output_dir, "ktp_export_results.xlsx")
        headers = ["Source File", "NIK", "Nama", "Tempat/Tgl Lahir", "Status Pernikahan", "Review Needed"]
        
        wb = Workbook()
        ws = wb.active
        ws.title = "KTP Data"
        
        ws.append(headers)
        
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        red_font = Font(color="9C0006")
        
        for row_idx, record in enumerate(all_ktp_records, start=2):
            row_data = [record.get(h, "") for h in headers]
            ws.append(row_data)
            
            if record.get("Review Needed", ""):
                review_cell = ws.cell(row=row_idx, column=6)
                review_cell.fill = red_fill
                review_cell.font = red_font
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[col_letter].width = adjusted_width

        wb.save(excel_filename)
            
        print(f"\n- Woo Yessir -")
        print(f"Exported {len(all_ktp_records)} KTP records to {excel_filename}")
    else:
        print("\nNo KTP data was extracted from the batch.")

    end_time = time.perf_counter()
    total_seconds = end_time - start_time
    hours, remainder = divmod(total_seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    print(f"Code took {int(hours)} hours, {int(minutes)} minutes, {seconds:.2f} seconds to execute.")

if __name__ == "__main__":
    main()